import streamlit as st
import openpyxl
import re
import io
from datetime import datetime
# Başlık
st.title("📄 ZTM003 > Yönlendirme Şablonu Aktarıcı")
# Dosya yükleme
uploaded_po = st.file_uploader("📂 Lütfen ZTM003 dosyasını (.xlsx) yükleyin:", type="xlsx")
uploaded_yon = st.file_uploader("📂 Lütfen Yönlendirme Şablon dosyasını (.xlsx) yükleyin:", type="xlsx")
# Normalizasyon fonksiyonu
def normalize(val):
   val = re.sub(r"[\r\n\t]", "", str(val)).lower()
   return re.sub(r"\s+", "", val)
# Mapping
mapping = {
   normalize("Alıcı"): "Sipariş veren bayi/dist Kodu",
   normalize("Üretim yeri"): "Yönlendirme Yapılan Fabrika Kodu (2. SN)",
   normalize("Kapı Çıkış Tarihi"): "Fatura Tarihi",
   normalize("Ürün"): "Ürün Kodu (SKU)",
   normalize("Teslimat Miktarı"): "Adet (Tava\\Koli\\Kasa)",
   normalize("yönlendirme nedeni"): "Yönlendirme yapma nedeni"
}
# Nakliye kod eşlemesi
nakliye_kod_map = {
   ("ZTIR", "Gidiş"): "ZTIR01",
   ("ZTIR", "Gidiş-Dönüş"): "ZTIR02",
   ("ZKMY", "Gidiş"): "ZKMY01",
   ("ZKMY", "Gidiş-Dönüş"): "ZKMY02"
}
# Dosyalar yüklendiyse devam et
if uploaded_po and uploaded_yon:
   wb_src = openpyxl.load_workbook(uploaded_po, data_only=True)
   ws_src = wb_src["Data"]
   wb_dst = openpyxl.load_workbook(uploaded_yon)
   ws_dst = wb_dst["Ana_sayfa"]
   # Başlık indeksleri
   src_headers = {normalize(cell.value): idx for idx, cell in enumerate(ws_src[1])}
   dst_headers = {normalize(cell.value): idx for idx, cell in enumerate(ws_dst[1])}
   # Gerekli indeksler
   sevk_idx = src_headers.get(normalize("Teslimat Miktarı"))
   nk_idx = src_headers.get(normalize("Nakliye araçları"))
   yon_idx = src_headers.get(normalize("Nakliye Tipi Tanımı"))
   dst_nt_idx = dst_headers.get(normalize("Nakliye Tipi"))
   # Aktarım başlıyor
   dst_row = 2
   aktarilan = 0
   for row in ws_src.iter_rows(min_row=2, values_only=True):
       sevk_miktar = row[sevk_idx] if sevk_idx is not None else None
       if sevk_miktar in [0, 0.0, "0", "0.0", None, ""]:
           continue
       # Sabit alanları aktar
       for src_col, dst_col in mapping.items():
           src_idx = src_headers.get(src_col)
           dst_idx = dst_headers.get(normalize(dst_col))
           if src_idx is None or dst_idx is None:
               continue
           ws_dst.cell(row=dst_row, column=dst_idx + 1, value=row[src_idx])
       # Nakliye kodunu hesapla
       nk_val = str(row[nk_idx]).strip().upper() if nk_idx is not None and row[nk_idx] else ""
       yon_val = str(row[yon_idx]).strip() if yon_idx is not None and row[yon_idx] else ""
       combined = nakliye_kod_map.get((nk_val, yon_val), f"{nk_val}{yon_val}")
       if dst_nt_idx is not None:
           ws_dst.cell(row=dst_row, column=dst_nt_idx + 1, value=combined)
       dst_row += 1
       aktarilan += 1
   # Dosyayı belleğe yaz
   output = io.BytesIO()
   wb_dst.save(output)
   output.seek(0)
   filename = f"Yönlendirme_{datetime.today().strftime('%d%m%Y')}.xlsx"
   # Sonuç
   st.success(f"✅ {aktarilan} satır aktarıldı.")
   st.download_button(
       label="📥 Aktarılan Dosyayı İndir",
       data=output,
       file_name=filename,
       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
   )
